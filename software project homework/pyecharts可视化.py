from pyecharts import options as opts
from pyecharts.charts import Bar
from pyecharts.commons.utils import JsCode
from pyecharts.globals import ThemeType
import openpyxl

if __name__ == "__main__":
    data = openpyxl.load_workbook('D:\软工作业——疫情统计\各省份疫情统计.xlsx')
    table = data['1']
    province = [
        "河北", "山西", "辽宁", "吉林", "黑龙江", "江苏",
        "浙江", "安徽", "福建", "江西", "山东", "河南", "湖北", "湖南",
        "广东", "海南", "四川", "贵州", "云南", "陕西", "北京", "天津", "上海", "重庆",
        "甘肃", "青海", "内蒙古", "广西", "西藏", "宁夏", "新疆"
    ]
    list_data = []
    for i in range(2, 9):
        list_data.append(table.cell(i, 1).value)
    print(list_data)

    day_all = [0, 0, 0, 0, 0, 0, 0]

    for i in range(2, 9):
        for j in range(0, 30):
            day_all[i-2] = day_all[i-2] + table.cell(i, j + 2).value
    print(day_all)
    list_all = []
    for i in range(0, 30):
        list_s = []
        for j in range(2, 9):
            list_s.append({"value": table.cell(j, i + 2).value, "percent": day_all[j - 2]})
        list_all.append(list_s)
    print(list_all)

    c = (
        Bar(init_opts=opts.InitOpts(theme=ThemeType.MACARONS))
        .add_xaxis(list_data)
        .add_yaxis(province[0], list_all[0], stack="stack1", category_gap="50%")
        .add_yaxis(province[1], list_all[1], stack="stack1", category_gap="50%")
        .add_yaxis(province[2], list_all[2], stack="stack1", category_gap="50%")
        .add_yaxis(province[3], list_all[3], stack="stack1", category_gap="50%")
        .add_yaxis(province[4], list_all[4], stack="stack1", category_gap="50%")
        .add_yaxis(province[5], list_all[5], stack="stack1", category_gap="50%")
        .add_yaxis(province[6], list_all[6], stack="stack1", category_gap="50%")
        .add_yaxis(province[7], list_all[7], stack="stack1", category_gap="50%")
        .add_yaxis(province[8], list_all[8], stack="stack1", category_gap="50%")
        .add_yaxis(province[9], list_all[9], stack="stack1", category_gap="50%")
        .add_yaxis(province[10], list_all[10], stack="stack1", category_gap="50%")
        .add_yaxis(province[11], list_all[11], stack="stack1", category_gap="50%")
        .add_yaxis(province[12], list_all[12], stack="stack1", category_gap="50%")
        .add_yaxis(province[13], list_all[13], stack="stack1", category_gap="50%")
        .add_yaxis(province[14], list_all[14], stack="stack1", category_gap="50%")
        .add_yaxis(province[15], list_all[15], stack="stack1", category_gap="50%")
        .add_yaxis(province[16], list_all[16], stack="stack1", category_gap="50%")
        .add_yaxis(province[17], list_all[17], stack="stack1", category_gap="50%")
        .add_yaxis(province[18], list_all[18], stack="stack1", category_gap="50%")
        .add_yaxis(province[19], list_all[19], stack="stack1", category_gap="50%")
        .add_yaxis(province[20], list_all[20], stack="stack1", category_gap="50%")
        .add_yaxis(province[21], list_all[21], stack="stack1", category_gap="50%")
        .add_yaxis(province[22], list_all[22], stack="stack1", category_gap="50%")
        .add_yaxis(province[23], list_all[23], stack="stack1", category_gap="50%")
        .add_yaxis(province[24], list_all[24], stack="stack1", category_gap="50%")
        .add_yaxis(province[25], list_all[25], stack="stack1", category_gap="50%")
        .add_yaxis(province[26], list_all[26], stack="stack1", category_gap="50%")
        .add_yaxis(province[27], list_all[27], stack="stack1", category_gap="50%")
        .add_yaxis(province[28], list_all[28], stack="stack1", category_gap="50%")
        .add_yaxis(province[29], list_all[29], stack="stack1", category_gap="50%")
        .set_series_opts(
            label_opts=opts.LabelOpts(is_show=False)
        )
        .set_global_opts(yaxis_opts=opts.AxisOpts(name='确诊人数'))
        .set_global_opts(xaxis_opts=opts.AxisOpts(name='日期'))
        .render("七日疫情各省及总体情况.html")
    )