import requests
import openpyxl
import time
import re
from lxml import etree
if __name__ == "__main__":
    # UA伪装
    headers = {
        "Accept": "image/avif,image/webp,*/*",
        "Accept-Encoding": "gzip, deflate",
        "Accept-Language": "zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2",
        "Connection": "keep-alive",
        "Referer":"http://www.nhc.gov.cn/",
        "Cookie":"yfx_c_g_u_id_10006654=_ck22090517092711557123132504572; yfx_f_l_v_t_10006654=f_t_1662368967158__r_t_1662514638127__v_t_1662514638127__r_c_2; yfx_mr_10006654=%3A%3Amarket_type_free_search%3A%3A%3A%3Abaidu%3A%3A%3A%3A%3A%3A%3A%3Awww.baidu.com%3A%3A%3A%3Apmf_from_free_search; yfx_mr_f_10006654=%3A%3Amarket_type_free_search%3A%3A%3A%3Abaidu%3A%3A%3A%3A%3A%3A%3A%3Awww.baidu.com%3A%3A%3A%3Apmf_from_free_search; yfx_key_10006654=; sVoELocvxVW0S=57yh5eHi6BlWwbYuOEUHFMNXf_2SF8UL5VWS1759zdOiiImwtyLuvBL1rWffIpGlLnMEMoxnpQBHoAej5Qug.gG; security_session_verify=fc465e4f1828940ec0438b63374ada0a; sVoELocvxVW0T=53SI0.DWUeQ7qqqDkmRH3_AToYARjKiHRH568jKOM4B.OPNB2axXw5kqAtweBhHBYQOYh3hRO8OaMl8SZuRBb4HDDy8wWx_H9KnDMJfOHJhLKqwvylr_gmnhMbVf7Xl1INInmRUZl8aTrrguv1MWZmyUOXCgg2aOx6_4J72Gm.uCLEdwxtjF7hWLGGpO..CyBUuKFNGN8o.f7i5cTf3DueMgKy959yMbnxH14vnDsH.wVdK4nQbz4PLAMqYCxYgwjT4eY2xujVIScYsPVxnC5uNL45UyizBvMCagu5cjSPGfoWPa5mHqCrzryZOZ96c0axMDYQTxAdJ7LeVECT_l6vDQTiBIU5G26AzdTJ07AoQza; insert_cookie=91349450",
        "Hosts":"www.nhc.gov.cn",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:104.0) Gecko/20100101 Firefox/104.0"
    }
    #获取卫健委疫情通报总页数
    home_url = "http://www.nhc.gov.cn/xcs/yqtb/list_gzbd.shtml"
    response = requests.get(url=home_url, headers=headers)
    time.sleep(0.5)
    page_text = response.text
    ex = "(?<='page_div',).*(?=, )"
    page_len = int(re.findall(ex, page_text, re.S)[0])
    #建立excel对象
    book = openpyxl.Workbook()
    sh = book.active
    sh.title = '疫情确诊及无症状感染人数统计'
    sh['A1'] = 'date'
    sh['B1'] = '新增确诊病例'
    sh['C1'] = '无症状感染者'
    row = 2
    #遍历卫健委疫情通报所有网页
    for length in range(1, page_len+1):
        url = format("http://www.nhc.gov.cn/xcs/yqtb/list_gzbd_%d.shtml" % length)
        if (length == 1):
            url = home_url
        response = requests.get(url=url, headers=headers)
        time.sleep(0.5)
        page_text = response.text
        tree = etree.HTML(page_text)
        #找到每日详情页链接
        li_list = tree.xpath('//div[@class="w1024 mb50"]/div[@class="list"]/ul/li')
        length = len(li_list)
        new_url = []
        for i in range(1, length+1):
            str1 = format('//div[@class="w1024 mb50"]/div[@class="list"]/ul/li[%d]/a/@href' % i)
            #拼接得到每日详情页链接
            new_url.append("http://www.nhc.gov.cn" + tree.xpath(str1)[0])
        for i in new_url:
            #获取每日详情页
            response = requests.get(url=i, headers=headers)
            time.sleep(0.5)
            #获取文本
            page_text = response.text
            tree = etree.HTML(page_text)
            text_list = tree.xpath('./body/div[@class="w1024 mb50"]/div[@class="list"]/div[@id="xw_box"][@class="con"]/p//text()')
            text = "".join(text_list)
            #正则匹配寻找相应信息
            ex1 = "(?<=新增确诊病例)\d+(?=例)"
            ex2 = "(?<=新增无症状感染者)\d+(?=例)"
            ex3 = "(?<=截至).*?(?=24时)"
            new = re.findall(ex1, text, re.S)
            asymptomatic = re.findall(ex2, text, re.S)
            date = re.findall(ex3, text, re.S)
            if not new:
                new = ""
            else:
                new = int(new[0])
            if not asymptomatic:
                asymptomatic = 0
            else:
                asymptomatic = int(asymptomatic[0])
            if not date:
                date = ''
            else:
                date = date[0]
            #将信息写入excel
            sh.cell(row, 1, date)
            sh.cell(row, 2, new)
            sh.cell(row, 3, asymptomatic)
            row += 1
            print('date:', end='')
            print(date)
            print('新增：', end='')
            print(new)
            print('无症状：', end='')
            print(asymptomatic)
    #保存excel
    book.save('疫情统计.xlsx')